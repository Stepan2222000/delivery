"""Excel I/O for parcels.

Layout (single sheet, frozen header, autofilter, dropdown on Статус):
    Трек | Статус | Заказан | Прибыл в США | Получен в США | Прибыл в КГ |
    Отправлен в РФ | Доставлен в РФ | Вес, кг | Заметка

Columns are read by name (not index), so order can be re-shuffled by the user.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

STATUS_RU = {
    "ordered": "заказано",
    "arrived_usa": "прибыл в США",
    "received_by_forwarder_usa": "получен в США",
    "in_shipment_usa_to_kg": "в пути в КГ",
    "arrived_kg": "прибыл в КГ",
    "in_shipment_kg_to_ru": "в пути в РФ",
    "delivered_ru": "доставлен в РФ",
    "not_received_ru": "не получено в РФ",
    "cancelled": "отменено",
}
STATUS_FROM_RU = {v: k for k, v in STATUS_RU.items()}

COLUMNS = [
    ("Трек",            "tracking_number"),
    ("Статус",          "status"),
    ("Заказан",         "ordered_at"),
    ("Прибыл в США",    "arrived_usa_at"),
    ("Получен в США",   "received_usa_at"),
    ("Прибыл в КГ",     "arrived_kg_at"),
    ("Отправлен в РФ",  "sent_ru_at"),  # virtual: shipments.sent_at via shipment_kg_to_ru_id
    ("Доставлен в РФ",  "delivered_ru_at"),
    ("Вес, кг",         "weight_kg"),
    ("Заметка",         "notes"),
]


def _date_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v)


def build_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """Build xlsx with frozen header, autofilter, dropdown on status column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Треки"

    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for r in rows:
        ws.append([
            r.get("tracking_number"),
            STATUS_RU.get(r.get("status"), r.get("status")),
            _date_str(r.get("ordered_at")),
            _date_str(r.get("arrived_usa_at")),
            _date_str(r.get("received_usa_at")),
            _date_str(r.get("arrived_kg_at")),
            _date_str(r.get("sent_ru_at")),
            _date_str(r.get("delivered_ru_at")),
            float(r["weight_kg"]) if r.get("weight_kg") is not None else None,
            r.get("notes"),
        ])

    # Dropdown for the status column.
    status_col_idx = next(i for i, (h, _) in enumerate(COLUMNS, start=1) if h == "Статус")
    col_letter = get_column_letter(status_col_idx)
    values_csv = ",".join(STATUS_RU.values())
    dv = DataValidation(type="list", formula1=f'"{values_csv}"', allow_blank=True)
    dv.error = "Допустимы только значения из списка"
    dv.errorTitle = "Неверный статус"
    ws.add_data_validation(dv)
    last_row = ws.max_row if ws.max_row > 1 else 1048576
    dv.add(f"{col_letter}2:{col_letter}{last_row}")

    # Autofilter + simple width tuning.
    ws.auto_filter.ref = ws.dimensions
    widths = {1: 26, 2: 22, 3: 12, 4: 14, 5: 14, 6: 14, 7: 16, 8: 16, 9: 10, 10: 40}
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def parse_xlsx(data: bytes) -> list[dict[str, Any]]:
    """Read xlsx into list of dicts keyed by internal field name.

    Returns one dict per data row. Empty rows skipped. Status names translated back
    to internal codes; weight coerced to float; dates parsed with tolerance.
    """
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []
        name_to_field = {label: field for label, field in COLUMNS}
        indexed_fields: list[str | None] = [
            name_to_field.get((str(h).strip() if h else "")) for h in header
        ]
        out: list[dict[str, Any]] = []
        for row in rows_iter:
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue
            rec: dict[str, Any] = {}
            for idx, value in enumerate(row):
                field = indexed_fields[idx] if idx < len(indexed_fields) else None
                if field is None:
                    continue
                if field == "status" and isinstance(value, str):
                    rec[field] = STATUS_FROM_RU.get(value.strip(), value.strip())
                elif field == "weight_kg" and value not in (None, ""):
                    try:
                        rec[field] = float(value)
                    except (TypeError, ValueError):
                        rec[field] = ("__invalid__", str(value))
                elif field == "tracking_number" and value is not None:
                    rec[field] = str(value).strip()
                elif value is None or (isinstance(value, str) and value.strip() == ""):
                    continue
                else:
                    rec[field] = value
            if "tracking_number" in rec:
                out.append(rec)
        return out
    finally:
        wb.close()
